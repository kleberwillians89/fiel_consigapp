from flask import Flask, render_template, request, redirect, session, url_for, send_file
import json
import csv
import openpyxl
from datetime import datetime
from io import StringIO, BytesIO
from openpyxl import Workbook

app = Flask(__name__)
app.secret_key = 'fiel-super-secreto'
LEADS_PATH = 'leads.json'

# ===================== PÁGINAS PÚBLICAS ========================

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/fgts', methods=['GET', 'POST'])
def fgts():
    if request.method == 'POST':
        dados = request.form.to_dict()
        salvar_lead(dados, 'FGTS')

        nome = dados.get('nome', '')
        cpf = dados.get('cpf', '')
        telefone = dados.get('telefone', '')
        email = dados.get('email', '')
        nascimento = dados.get('nascimento', '')
        cidade = dados.get('cidade', '')
        valor = dados.get('valor', '')

        msg = f"""Olá! Gostaria de antecipar meu FGTS com a Fiel Crédito.
📌 *Nome:* {nome}
📌 *CPF:* {cpf}
📌 *WhatsApp:* {telefone}
📌 *E-mail:* {email}
📌 *Nascimento:* {nascimento}
📌 *Cidade:* {cidade}
📌 *Valor solicitado:* R$ {valor}"""

        msg = msg.replace("\n", "%0A")
        link = f"https://wa.me/5511986531008?text={msg}"
        return redirect(link)

    return render_template('fgts.html')

@app.route('/portabilidade', methods=['GET', 'POST'])
def portabilidade():
    if request.method == 'POST':
        salvar_lead(request.form.to_dict(), 'Portabilidade INSS')
        return redirect(url_for('confirmacao_portabilidade'))
    return render_template('portabilidade.html')

@app.route('/clt', methods=['GET', 'POST'])
def clt():
    if request.method == 'POST':
        dados = request.form.to_dict()
        salvar_lead(dados, 'Empréstimo CLT')

        nome = dados.get('nome', '')
        cpf = dados.get('cpf', '')
        telefone = dados.get('telefone', '')
        empresa = dados.get('empresa', '')
        salario = dados.get('salario', '')
        tempo = dados.get('tempo_empresa', '')

        msg = f"""Olá! Gostaria de fazer uma simulação de crédito CLT com a Fiel Crédito.
📌 *Nome:* {nome}
📌 *CPF:* {cpf}
📌 *Telefone:* {telefone}
📌 *Empresa:* {empresa}
📌 *Último salário:* R$ {salario}
📌 *Tempo de empresa:* {tempo} meses"""

        msg = msg.replace("\n", "%0A")
        link = f"https://wa.me/5511986531008?text={msg}"
        return redirect(link)

    return render_template('clt.html')

@app.route('/inss', methods=['GET', 'POST'])
def inss():
    if request.method == 'POST':
        dados = request.form.to_dict()
        salvar_lead(dados, 'Empréstimo INSS')

        nome = dados.get('nome', '')
        cpf = dados.get('cpf', '')
        telefone = dados.get('telefone', '')
        valor = dados.get('valor', '')
        parcela = dados.get('parcela', '')

        msg = f"""Olá! Gostaria de contratar um empréstimo INSS com a Fiel Crédito.
📌 *Nome:* {nome}
📌 *CPF:* {cpf}
📌 *WhatsApp:* {telefone}
📌 *Valor desejado:* R$ {valor}
📌 *Parcela pretendida:* R$ {parcela}
📌 *Juros:* 1,80% a.m."""

        msg = msg.replace("\n", "%0A")
        link = f"https://wa.me/5511986531008?text={msg}"
        return redirect(link)

    return render_template('inss.html')

@app.route('/serasa', methods=['GET', 'POST'])
def serasa():
    if request.method == 'POST':
        dados = request.form.to_dict()
        salvar_lead(dados, 'Serasa Limpa Nome')

        nome = dados.get('nome', '')
        cpf = dados.get('cpf', '')
        telefone = dados.get('telefone', '')
        govbr = dados.get('govbr', '')
        divida = dados.get('divida', '')
        score = dados.get('score', '')

        msg = f"""Olá! Gostaria de limpar meu nome com a Fiel Crédito.
📌 *Nome:* {nome}
📌 *CPF:* {cpf}
📌 *WhatsApp:* {telefone}
📌 *Tem acesso ao gov.br?* {govbr}
📌 *Valor da dívida:* R$ {divida}
📌 *Score atual:* {score}"""

        msg = msg.replace("\n", "%0A")
        link = f"https://wa.me/5511986531008?text={msg}"
        return redirect(link)

    return render_template('serasa.html')

@app.route('/chatbot', methods=['GET', 'POST'])
def chatbot():
    if request.method == 'POST':
        dados = request.form.to_dict()
        salvar_lead(dados, 'Chatbot')
        nome = dados.get('nome', '')
        cpf = dados.get('cpf', '')
        telefone = dados.get('telefone', '')
        cidade = dados.get('cidade', '')
        motivo = dados.get('motivo', '')

        msg = f"""Olá! Acabei de preencher o formulário de atendimento da Fiel Crédito.
📌 *Nome:* {nome}
📌 *CPF:* {cpf}
📌 *WhatsApp:* {telefone}
📌 *Cidade:* {cidade}
📌 *Motivo do contato:* {motivo}"""

        msg = msg.replace("\n", "%0A")
        link = f"https://wa.me/5511986531008?text={msg}"
        return redirect(link)

    return render_template('chatbot.html')



# ===================== CONFIRMAÇÕES ========================

@app.route('/confirmacao')
def confirmacao():
    return render_template('confirmacao.html')

@app.route('/confirmacao_portabilidade')
def confirmacao_portabilidade():
    return render_template('confirmacao_portabilidade.html')

@app.route('/confirmacao_clt')
def confirmacao_clt():
    return render_template('confirmacao_clt.html')

@app.route('/confirmacao_inss')
def confirmacao_inss():
    return render_template('confirmacao_inss.html')

@app.route('/confirmacao_serasa')
def confirmacao_serasa():
    return render_template('confirmacao_serasa.html')

# ===================== PAINEL ADMINISTRATIVO ========================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        if request.form['senha'] == 'admin123':
            session['logado'] = True
            return redirect(url_for('dashboard'))
        else:
            return 'Senha incorreta'
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
def dashboard():
    if not session.get('logado'):
        return redirect(url_for('login'))
    leads = carregar_leads()
    return render_template('dashboard.html', leads=leads)

@app.route('/importados')
def importados():
    if not session.get('logado'):
        return redirect(url_for('login'))
    try:
        with open("importados.json", "r") as f:
            lista = json.load(f)
    except:
        lista = []
    return render_template("importados.html", leads=lista)

@app.route('/importar', methods=['GET', 'POST'])
def importar():
    if not session.get('logado'):
        return redirect(url_for('login'))

    if request.method == 'POST':
        arquivo = request.files['arquivo']
        nome_arquivo = arquivo.filename

        lista = []

        if nome_arquivo.endswith('.csv'):
            conteudo = arquivo.read().decode('utf-8').splitlines()
            leitor = csv.DictReader(conteudo)
            for linha in leitor:
                lista.append({
                    "nome": linha.get("nome", ""),
                    "telefone": linha.get("telefone", ""),
                    "cpf": linha.get("cpf", ""),
                    "servico": linha.get("produto", "")
                })

        elif nome_arquivo.endswith('.xlsx'):
            planilha = openpyxl.load_workbook(arquivo)
            aba = planilha.active
            for i, linha in enumerate(aba.iter_rows(min_row=2, values_only=True)):
                lista.append({
                    "nome": linha[0],
                    "telefone": str(linha[1]),
                    "cpf": str(linha[2]),
                    "servico": linha[3]
                })

        with open("importados.json", "w") as f:
            json.dump(lista, f, indent=2)

        return redirect(url_for('importados'))

    return render_template("importar.html")

@app.route('/exportar-xlsx')
def exportar_xlsx():
    if not session.get('logado'):
        return redirect(url_for('login'))

    try:
        with open(LEADS_PATH, 'r') as f:
            leads = json.load(f)
    except:
        leads = []

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    headers = ["Data", "Serviço", "Nome", "CPF", "Telefone", "E-mail", "Nascimento", "Cidade", "Valor"]
    ws.append(headers)

    for lead in leads:
        ws.append([
            lead.get("data", ""),
            lead.get("servico", ""),
            lead.get("nome", ""),
            lead.get("cpf", ""),
            lead.get("telefone", ""),
            lead.get("email", ""),
            lead.get("nascimento", ""),
            lead.get("cidade", ""),
            lead.get("valor", "")
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name='leads_fiel.xlsx',
        as_attachment=True
    )

# ===================== FUNÇÕES AUXILIARES ========================

def salvar_lead(formulario, servico):
    formulario['data'] = datetime.now().strftime("%Y-%m-%d %H:%M")
    formulario['servico'] = servico
    leads = carregar_leads()
    leads.append(formulario)
    with open(LEADS_PATH, 'w') as f:
        json.dump(leads, f, indent=2)

def carregar_leads():
    try:
        with open(LEADS_PATH, 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return []

# ===================== RODAR APP ========================

import os

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)
